import json

from openpyxl import load_workbook

from app.config import settings
from app.models import Dataset, DatasetRow, User
from app.services.dataset_store import (
    create_dataset_from_upload,
    read_dataset_range,
    write_csv_export,
    write_jsonl_export,
    write_xlsx_export,
)


async def _user(session_factory, username="store_user") -> int:
    async with session_factory() as s:
        user = User(username=username, display_name="x")
        s.add(user)
        await s.commit()
        return user.id


async def test_import_csv_to_shards_visible_rows(tmp_path, session_factory):
    uid = await _user(session_factory)
    src = tmp_path / "big.csv"
    src.write_text("q,a\nq1,a1\nq2,a2\nq3,a3\n", encoding="utf-8")
    async with session_factory() as s:
        datasets = await create_dataset_from_upload(
            s, user_id=uid, filename="big.csv", source_path=src,
            data_dir=settings.data_dir, shard_size=2)
        ds = datasets[0]
        assert ds.row_count == 3
        assert ds.total_rows_including_header == 4
        assert ds.header_row == 1 and ds.data_start_row == 2
        manifest = json.loads(ds.manifest_json)
        assert [sh["start_file_row"] for sh in manifest["shards"]] == [2, 4]

        page = await read_dataset_range(
            s, ds, data_dir=settings.data_dir, start_row=1, end_row=3)
    assert page["rows"][0] == {"__row_type": "header", "columns": ["q", "a"]}
    assert page["rows"][1:] == [{"q": "q1", "a": "a1"}, {"q": "q2", "a": "a2"}]


async def test_import_jsonl_has_no_header(tmp_path, session_factory):
    uid = await _user(session_factory)
    src = tmp_path / "seed.jsonl"
    src.write_text('{"q":"q1","a":"a1"}\n{"q":"q2","a":"a2"}\n', encoding="utf-8")
    async with session_factory() as s:
        ds = (await create_dataset_from_upload(
            s, user_id=uid, filename="seed.jsonl", source_path=src,
            data_dir=settings.data_dir, shard_size=1))[0]
        assert ds.header_row is None and ds.data_start_row == 1
        page = await read_dataset_range(
            s, ds, data_dir=settings.data_dir, start_row=1, end_row=1)
    assert page["rows"] == [{"q": "q1", "a": "a1"}]


async def test_read_range_projects_columns_and_truncates(tmp_path, session_factory):
    uid = await _user(session_factory)
    src = tmp_path / "wide.csv"
    src.write_text("q,a,b\nq1,a1,b1\nq2,a2,b2\nq3,a3,b3\n", encoding="utf-8")
    async with session_factory() as s:
        ds = (await create_dataset_from_upload(
            s, user_id=uid, filename="wide.csv", source_path=src,
            data_dir=settings.data_dir, shard_size=2))[0]
        page = await read_dataset_range(
            s, ds, data_dir=settings.data_dir, start_row=2, end_row=4,
            columns=["q"], max_rows=2)
    assert page["columns"] == ["q"]
    assert page["rows"] == [{"q": "q1"}, {"q": "q2"}]
    assert page["truncated"] is True


async def test_lazy_migrates_legacy_dataset_rows(tmp_path, session_factory):
    uid = await _user(session_factory)
    async with session_factory() as s:
        ds = Dataset(user_id=uid, name="legacy", original_filename="legacy.jsonl",
                     row_count=2, columns_json='["q"]')
        s.add(ds)
        await s.commit()
        s.add_all([
            DatasetRow(dataset_id=ds.id, idx=0, data_json='{"q":"q1"}'),
            DatasetRow(dataset_id=ds.id, idx=1, data_json='{"q":"q2"}'),
        ])
        await s.commit()
        ds_id = ds.id

    async with session_factory() as s:
        ds = await s.get(Dataset, ds_id)
        page = await read_dataset_range(
            s, ds, data_dir=settings.data_dir, start_row=1, end_row=2)
        ds = await s.get(Dataset, ds_id)
        assert json.loads(ds.manifest_json)["shards"]
    assert page["rows"] == [{"q": "q1"}, {"q": "q2"}]


async def test_stream_csv_and_jsonl_exports_from_shards(tmp_path, session_factory):
    uid = await _user(session_factory)
    src = tmp_path / "export.csv"
    src.write_text("q,a\nq1,a1\nq2,a2\n", encoding="utf-8")
    async with session_factory() as s:
        ds = (await create_dataset_from_upload(
            s, user_id=uid, filename="export.csv", source_path=src,
            data_dir=settings.data_dir, shard_size=10))[0]
        csv_path = await write_csv_export(s, ds, settings.data_dir, tmp_path / "out.csv")
        jsonl_path = await write_jsonl_export(s, ds, settings.data_dir, tmp_path / "out.jsonl")

    assert csv_path.read_text(encoding="utf-8").splitlines() == ["q,a", "q1,a1", "q2,a2"]
    assert [json.loads(line) for line in jsonl_path.read_text(encoding="utf-8").splitlines()] == [
        {"q": "q1", "a": "a1"}, {"q": "q2", "a": "a2"}]


async def test_export_xlsx_splits_sheets(monkeypatch, tmp_path, session_factory):
    from app.services import dataset_store

    monkeypatch.setattr(dataset_store, "EXCEL_MAX_DATA_ROWS_PER_SHEET", 2)
    uid = await _user(session_factory)
    src = tmp_path / "book.csv"
    src.write_text("q\nq1\nq2\nq3\n", encoding="utf-8")
    async with session_factory() as s:
        ds = (await create_dataset_from_upload(
            s, user_id=uid, filename="book.csv", source_path=src,
            data_dir=settings.data_dir, shard_size=10))[0]
        path = await write_xlsx_export(s, ds, settings.data_dir, tmp_path / "out.xlsx")

    wb = load_workbook(path, read_only=True)
    assert wb.sheetnames == ["data", "data_2"]
    assert [row for row in wb["data"].iter_rows(values_only=True)] == [("q",), ("q1",), ("q2",)]
    assert [row for row in wb["data_2"].iter_rows(values_only=True)] == [("q",), ("q3",)]
