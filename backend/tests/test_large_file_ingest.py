"""Spec 1（大文件统一摄入/导出）回归测试。
Task 1: 基础设施（紧凑序列化 / import_error 列）。后续任务陆续追加。
"""
import asyncio
import datetime
import io
import json

from openpyxl import Workbook

from app.models import Dataset, User
from app.services.dataset_store import _dumps_row


def _xlsx(path, sheets):
    wb = Workbook()
    wb.remove(wb.active)
    for title, rows in sheets:
        ws = wb.create_sheet(title)
        for r in rows:
            ws.append(r)
    wb.save(path)


async def _user(session_factory, username) -> int:
    async with session_factory() as s:
        u = User(username=username, display_name="x")
        s.add(u)
        await s.commit()
        return u.id


# --- Task 1: _dumps_row 紧凑分隔符 -----------------------------------------

def test_dumps_row_uses_compact_separators():
    assert _dumps_row({"a": 1, "b": "x"}) == '{"a":1,"b":"x"}'   # 无逗号/冒号后空格


def test_dumps_row_compact_still_neutralizes_and_serializes():
    out = _dumps_row({"x": float("inf"), "dt": datetime.datetime(2024, 1, 2, 3, 4)})
    parsed = json.loads(out)
    assert parsed["x"] is None and "2024-01-02" in parsed["dt"]
    assert ", " not in out and ": " not in out


# --- Task 1: Dataset.import_error 列 ----------------------------------------

async def test_dataset_has_import_error_column(session_factory):
    async with session_factory() as s:
        u = User(username="ingest_col_user", display_name="x")
        s.add(u)
        await s.commit()
        ds = Dataset(user_id=u.id, name="d", status="failed", import_error="boom")
        s.add(ds)
        await s.commit()
        got = await s.get(Dataset, ds.id)
    assert got.import_error == "boom"


# --- Task 2: 结构探测 / 行迭代 / 批量写分片 -------------------------------

def test_detect_csv_structure(tmp_path):
    p = tmp_path / "d.csv"
    p.write_text("q,a\n1,2\n", encoding="utf-8")
    from app.services.dataset_store import detect_upload_structure
    units = detect_upload_structure("d.csv", p)
    assert len(units) == 1
    u = units[0]
    assert u.columns == ["q", "a"] and u.header_row == 1 and u.data_start_row == 2
    assert u.reader == "csv" and u.original_format == "csv" and u.name == "d"


def test_detect_jsonl_structure(tmp_path):
    p = tmp_path / "d.jsonl"
    p.write_text('{"a":1}\n', encoding="utf-8")
    from app.services.dataset_store import detect_upload_structure
    units = detect_upload_structure("d.jsonl", p)
    assert len(units) == 1 and units[0].columns == [] and units[0].reader == "jsonl"
    assert units[0].header_row is None and units[0].data_start_row == 1


def test_detect_json_uses_json_reader(tmp_path):
    p = tmp_path / "d.json"
    p.write_text('[{"a":1}]', encoding="utf-8")
    from app.services.dataset_store import detect_upload_structure
    u = detect_upload_structure("d.json", p)[0]
    assert u.reader == "json" and u.original_format == "jsonl"


def test_detect_multisheet_xlsx(tmp_path):
    p = tmp_path / "book.xlsx"
    _xlsx(p, [("alpha", [["q", "a"], ["1", "2"]]), ("beta", [["x"], ["9"]])])
    from app.services.dataset_store import detect_upload_structure
    units = detect_upload_structure("book.xlsx", p)
    assert [u.name for u in units] == ["book-alpha", "book-beta"]
    assert units[0].columns == ["q", "a"] and units[0].sheet_index == 0
    assert units[1].columns == ["x"] and units[1].sheet_index == 1


def test_detect_xlsx_skips_dataless_sheet(tmp_path):
    p = tmp_path / "b.xlsx"
    _xlsx(p, [("hasdata", [["c"], ["v"]]), ("headeronly", [["h"]])])
    from app.services.dataset_store import detect_upload_structure
    units = detect_upload_structure("b.xlsx", p)
    assert len(units) == 1 and units[0].name == "b" and units[0].sheet_index == 0


def test_rows_for_unit_csv(tmp_path):
    p = tmp_path / "d.csv"
    p.write_text("q,a\n1,2\n3,4\n", encoding="utf-8")
    from app.services.dataset_store import detect_upload_structure, rows_for_unit
    u = detect_upload_structure("d.csv", p)[0]
    assert list(rows_for_unit(u, p)) == [{"q": "1", "a": "2"}, {"q": "3", "a": "4"}]


def test_rows_for_unit_xlsx_sheet(tmp_path):
    p = tmp_path / "b.xlsx"
    _xlsx(p, [("s1", [["q", "a"], ["1", "2"]]), ("s2", [["x"], ["9"]])])
    from app.services.dataset_store import detect_upload_structure, rows_for_unit
    units = detect_upload_structure("b.xlsx", p)
    assert list(rows_for_unit(units[1], p)) == [{"x": "9"}]


def test_parse_and_write_shards_with_progress(tmp_path):
    p = tmp_path / "d.csv"
    p.write_text("q\n" + "\n".join(str(i) for i in range(5)) + "\n", encoding="utf-8")
    from app.services.dataset_store import detect_upload_structure, parse_and_write_shards
    u = detect_upload_structure("d.csv", p)[0]
    seen = []
    manifest, cols, n = parse_and_write_shards(
        source_path=p, unit=u, data_dir=tmp_path / "dd",
        user_id=1, dataset_id=1, version=1, shard_size=2, progress_cb=seen.append)
    assert n == 5 and cols == ["q"]
    assert len(manifest["shards"]) == 3          # 2+2+1
    assert seen == [2, 4, 5]                      # 每分片关闭回调一次累计行数


# --- Task 3: IngestManager 后台摄入 + resume ------------------------------

from app.config import settings  # noqa: E402
from app.services.dataset_store import ParseUnit, dataset_root, detect_upload_structure  # noqa: E402
from app.services.ingest_manager import ingest_manager, resume_unfinished  # noqa: E402


async def _placeholder(session_factory, uid, name, src, unit) -> int:
    async with session_factory() as s:
        ds = Dataset(
            user_id=uid, name=name, source="upload", original_filename=src.name,
            original_format=unit.original_format, file_path=str(src), status="importing",
            header_row=unit.header_row, data_start_row=unit.data_start_row,
            columns_json=json.dumps(unit.columns, ensure_ascii=False))
        s.add(ds)
        await s.commit()
        return ds.id


async def test_ingest_success_marks_ready_and_reclaims_source(session_factory):
    uid = await _user(session_factory, "ingest_ok")
    src = settings.data_dir / "uploads" / "ingest_ok.csv"
    src.parent.mkdir(parents=True, exist_ok=True)
    src.write_text("q,a\n1,2\n3,4\n", encoding="utf-8")
    unit = detect_upload_structure("ingest_ok.csv", src)[0]
    ds_id = await _placeholder(session_factory, uid, "ingest_ok", src, unit)
    await ingest_manager._run_ingest(ds_id, src, unit, 1, uid, session_factory)
    async with session_factory() as s:
        got = await s.get(Dataset, ds_id)
    assert got.status == "ready" and got.row_count == 2 and got.import_error == ""
    assert list(dataset_root(settings.data_dir, uid, ds_id, 1).glob("part-*.jsonl"))
    assert not src.exists()                       # 源文件解析后回收


async def test_ingest_failure_marks_failed_and_cleans_orphans(session_factory):
    uid = await _user(session_factory, "ingest_fail")
    src = settings.data_dir / "uploads" / "bad.xlsx"
    src.parent.mkdir(parents=True, exist_ok=True)
    src.write_bytes(b"not a real workbook")
    unit = ParseUnit(name="bad", columns=["x"], header_row=1, data_start_row=2,
                     original_format="xlsx", reader="xlsx", sheet_index=0)
    ds_id = await _placeholder(session_factory, uid, "bad", src, unit)
    await ingest_manager._run_ingest(ds_id, src, unit, 1, uid, session_factory)
    async with session_factory() as s:
        got = await s.get(Dataset, ds_id)
    assert got.status == "failed" and got.import_error
    assert not dataset_root(settings.data_dir, uid, ds_id, 1).parent.exists()  # 孤儿分片清理


async def test_ingest_submit_runs_to_ready(session_factory):
    uid = await _user(session_factory, "ingest_submit")
    src = settings.data_dir / "uploads" / "sub.csv"
    src.parent.mkdir(parents=True, exist_ok=True)
    src.write_text("q\n1\n2\n", encoding="utf-8")
    unit = detect_upload_structure("sub.csv", src)[0]
    ds_id = await _placeholder(session_factory, uid, "sub", src, unit)
    ingest_manager.submit(ds_id, source_path=src, unit=unit, version=1,
                          user_id=uid, session_factory=session_factory)
    await ingest_manager._running[ds_id]          # 等后台任务完成
    async with session_factory() as s:
        got = await s.get(Dataset, ds_id)
    assert got.status == "ready" and got.row_count == 2


async def test_resume_unfinished_fails_stale_importing(session_factory):
    uid = await _user(session_factory, "resume_u")
    src = settings.data_dir / "uploads" / "stale.csv"
    src.parent.mkdir(parents=True, exist_ok=True)
    src.write_text("q\n1\n", encoding="utf-8")
    async with session_factory() as s:
        ds = Dataset(user_id=uid, name="stale", status="importing", file_path=str(src),
                     original_filename="stale.csv", original_format="csv")
        s.add(ds)
        await s.commit()
        ds_id = ds.id
    root = dataset_root(settings.data_dir, uid, ds_id, 1)
    root.mkdir(parents=True, exist_ok=True)
    (root / "part-000001.jsonl").write_text("{}\n", encoding="utf-8")
    n = await resume_unfinished(session_factory)
    assert n >= 1
    async with session_factory() as s:
        got = await s.get(Dataset, ds_id)
    assert got.status == "failed" and "重启" in got.import_error
    assert not root.parent.exists() and not src.exists()


# --- Task 4: /upload 端点 后台摄入 ----------------------------------------

from conftest import upload_ready, wait_ready  # noqa: E402


async def test_upload_endpoint_returns_importing_then_ready(auth_client):
    r = await auth_client.post(
        "/api/datasets/upload",
        files=[("files", ("u.csv", b"q,a\n1,2\n3,4\n", "application/octet-stream"))])
    assert r.status_code == 200
    ph = r.json()[0]
    assert ph["status"] == "importing" and ph["row_count"] == 0   # 端点秒回占位
    ds = await wait_ready(auth_client, ph["id"])
    assert ds["status"] == "ready" and ds["row_count"] == 2
    # 源文件解析后回收
    assert not list((settings.data_dir / "uploads").rglob("*.csv"))


async def test_upload_endpoint_deep_parse_error_becomes_failed(auth_client):
    # jsonl 首行合法(探测通过 importing)，第二行畸形 → 后台摄入失败 → status=failed
    r = await auth_client.post(
        "/api/datasets/upload",
        files=[("files", ("bad.jsonl", b'{"a":1}\n{bad}\n', "application/octet-stream"))])
    assert r.status_code == 200
    ds_id = r.json()[0]["id"]
    for _ in range(400):
        body = (await auth_client.get(f"/api/datasets/{ds_id}")).json()
        if body["status"] != "importing":
            break
        await asyncio.sleep(0.01)
    assert body["status"] == "failed" and body["import_error"]


async def test_upload_endpoint_excel_size_gate(auth_client, monkeypatch):
    monkeypatch.setattr(settings, "max_excel_upload_bytes", 10)
    buf = io.BytesIO()
    wb = Workbook()
    wb.active.append(["a", "b"])
    wb.active.append([1, 2])
    wb.save(buf)
    r = await auth_client.post(
        "/api/datasets/upload",
        files=[("files", ("big.xlsx", buf.getvalue(), "application/octet-stream"))])
    assert r.status_code == 422 and "Excel" in r.json()["detail"]


async def test_upload_endpoint_unsupported_extension(auth_client):
    r = await auth_client.post(
        "/api/datasets/upload",
        files=[("files", ("x.pdf", b"%PDF-1.4", "application/octet-stream"))])
    assert r.status_code == 422
