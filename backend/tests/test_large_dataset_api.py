from io import BytesIO

from conftest import wait_ready
from openpyxl import load_workbook


async def _upload(client, name: str, content: bytes):
    """上传并等后台摄入到 ready，返回 ready 数据集 dict 列表。"""
    r = await client.post(
        "/api/datasets/upload",
        files=[("files", (name, content, "application/octet-stream"))],
    )
    assert r.status_code == 200
    return [await wait_ready(client, d["id"]) for d in r.json()]


async def test_upload_csv_returns_large_dataset_metadata(auth_client):
    ds = (await _upload(auth_client, "people.csv", b"name,age\nAda,36\nLinus,55\n"))[0]
    assert ds["name"] == "people"
    assert ds["row_count"] == 2
    assert ds["total_rows_including_header"] == 3
    assert ds["status"] == "ready"
    assert ds["imported_rows"] == 2
    assert ds["original_format"] == "csv"
    assert ds["version"] == 1
    assert ds["header_row"] == 1
    assert ds["data_start_row"] == 2
    assert ds["columns"] == ["name", "age"]


async def test_rows_start_end_include_csv_header(auth_client):
    ds = (await _upload(auth_client, "people.csv", b"name,age\nAda,36\nLinus,55\n"))[0]
    r = await auth_client.get(f"/api/datasets/{ds['id']}/rows?start_row=1&end_row=2")
    assert r.status_code == 200
    payload = r.json()
    assert payload["total_rows_including_header"] == 3
    assert payload["header_row"] == 1
    assert payload["data_start_row"] == 2
    assert payload["rows"] == [
        {"__row_type": "header", "columns": ["name", "age"]},
        {"name": "Ada", "age": "36"},
    ]


async def test_rows_jsonl_first_row_is_record(auth_client):
    ds = (await _upload(auth_client, "items.jsonl", b'{"id": 1}\n{"id": 2}\n'))[0]
    r = await auth_client.get(f"/api/datasets/{ds['id']}/rows?start_row=1&end_row=1")
    assert r.status_code == 200
    payload = r.json()
    assert payload["header_row"] is None
    assert payload["data_start_row"] == 1
    assert payload["rows"] == [{"id": 1}]


async def test_rows_column_projection(auth_client):
    ds = (await _upload(auth_client, "people.csv", b"name,age,city\nAda,36,London\n"))[0]
    r = await auth_client.get(
        f"/api/datasets/{ds['id']}/rows?start_row=1&end_row=2&columns=name,city"
    )
    assert r.status_code == 200
    assert r.json()["rows"] == [
        {"__row_type": "header", "columns": ["name", "city"]},
        {"name": "Ada", "city": "London"},
    ]


async def test_export_original_csv_streams(auth_client):
    ds = (await _upload(auth_client, "people.csv", b"name,age\nAda,36\n"))[0]
    r = await auth_client.get(f"/api/datasets/{ds['id']}/export?format=original")
    assert r.status_code == 200
    assert "text/csv" in r.headers["content-type"]
    assert r.text.splitlines() == ["name,age", "Ada,36"]


async def test_export_xlsx_uses_write_only_workbook(auth_client):
    ds = (await _upload(auth_client, "people.csv", b"name,age\nAda,36\n"))[0]
    r = await auth_client.get(f"/api/datasets/{ds['id']}/export?format=xlsx")
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.content), read_only=True, data_only=True)
    ws = wb.active
    assert list(ws.iter_rows(values_only=True)) == [("name", "age"), ("Ada", "36")]
