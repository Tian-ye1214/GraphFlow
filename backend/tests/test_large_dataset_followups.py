"""Regression tests for the 6 must-fix review findings on the large-dataset feature:
1 删除回收分片磁盘 / 2 损坏 Excel→422 / 3 超长 CSV 单元格不 500 /
4 CRUD 非有限数中和 / 5 嵌套单元格导出 xlsx 不 500 / 6 分页不被 agent 预算截断。

外加复审（活体完备性 Workflow）追出的后续缺口回归：
A xlsx 导出含 XML 控制字符/超大整数不 500 / B xlsx 上传含 datetime 不 500 /
C 数据集导出文件回收 / D csv 嵌套单元格串成 JSON / E 范围读表头伪行不占行数顶。
"""
import csv as _csv
import datetime
import json
from io import BytesIO, StringIO

from openpyxl import Workbook, load_workbook

from conftest import upload_ready

from app.config import settings
from app.models import User
from app.services.dataset_store import create_dataset_from_upload, read_dataset_range


async def _upload(client, name: str, content: bytes):
    # 原始上传响应（importing 占位或 422）；需 ready 数据的用例改用 conftest.upload_ready。
    return await client.post(
        "/api/datasets/upload",
        files=[("files", (name, content, "application/octet-stream"))],
    )


async def _user(session_factory, username="followup_user") -> int:
    async with session_factory() as s:
        user = User(username=username, display_name="x")
        s.add(user)
        await s.commit()
        return user.id


# --- Fix 1: delete reclaims shard disk -------------------------------------

async def test_delete_dataset_reclaims_shard_files(auth_client):
    ds = (await upload_ready(auth_client, "p.csv", b"q,a\nq1,a1\nq2,a2\n"))[0]
    assert list((settings.data_dir / "datasets").rglob("part-*.jsonl"))
    r = await auth_client.delete(f"/api/datasets/{ds['id']}")
    assert r.status_code == 200
    assert not list((settings.data_dir / "datasets").rglob("part-*.jsonl"))


async def test_unlink_run_exports_removes_run_artifact_dir():
    from app.services.run_service import unlink_run_exports

    run_dir = settings.data_dir / "runs" / "777" / "node_x"
    run_dir.mkdir(parents=True, exist_ok=True)
    (run_dir / "part-000001.jsonl").write_text("{}\n", encoding="utf-8")
    unlink_run_exports([777], settings.data_dir)
    assert not (settings.data_dir / "runs" / "777").exists()


# --- Fix 2: corrupt Excel upload -> 422, not 500 ---------------------------

async def test_corrupt_xlsx_upload_returns_422(auth_client):
    r = await _upload(auth_client, "bad.xlsx", b"this is not a real workbook")
    assert r.status_code == 422


def _truncated_sheet_xlsx() -> bytes:
    """合法 zip、但 sheet XML 截断：load_workbook 过得了，惰性 iter_rows 抛 XML ParseError。"""
    import zipfile

    from openpyxl import Workbook

    buf = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.append(["a", "b"])
    ws.append([1, 2])
    wb.save(buf)
    src = zipfile.ZipFile(BytesIO(buf.getvalue()))
    out = BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as z:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename.endswith("sheet1.xml"):
                data = data[: len(data) // 2]      # 截一半 → 未闭合标签
            z.writestr(item, data)
    return out.getvalue()


async def test_truncated_sheet_xml_xlsx_returns_422(auth_client):
    r = await _upload(auth_client, "broken.xlsx", _truncated_sheet_xlsx())
    assert r.status_code == 422


# --- Fix 3: a single >128KB CSV cell is accepted (legit large-text row) -----

async def test_large_csv_cell_uploads(auth_client):
    big = "x" * 200_000
    content = f"q,a\nq1,{big}\n".encode("utf-8")
    ds = (await upload_ready(auth_client, "big.csv", content))[0]
    assert ds["row_count"] == 1


# --- Fix 4: CRUD non-finite values are neutralized to null in canonical shards

async def test_crud_constant_non_finite_neutralized(tmp_path, session_factory):
    from app.services.dataset_crud import apply_dataset_operations

    uid = await _user(session_factory)
    src = tmp_path / "d.csv"
    src.write_text("q\nq1\n", encoding="utf-8")
    async with session_factory() as s:
        ds = (await create_dataset_from_upload(
            s, user_id=uid, filename="d.csv", source_path=src,
            data_dir=settings.data_dir, shard_size=10))[0]
        new = await apply_dataset_operations(
            s, source=ds, user_id=uid, data_dir=settings.data_dir,
            operations=[{"op": "add_constant_column", "name": "c", "value": float("inf")}])
        page = await read_dataset_range(
            s, new, data_dir=settings.data_dir, start_row=2, end_row=2)
    assert page["rows"] == [{"q": "q1", "c": None}]
    for p in (settings.data_dir / "datasets").rglob("part-*.jsonl"):
        text = p.read_text(encoding="utf-8")
        assert "Infinity" not in text and "NaN" not in text


# --- Fix 5: xlsx export of nested (dict/list) cells stringifies, not 500 ----

async def test_export_xlsx_stringifies_nested_cells(auth_client):
    ds = (await upload_ready(auth_client, "n.jsonl", b'{"a": {"x": 1}}\n'))[0]
    r = await auth_client.get(f"/api/datasets/{ds['id']}/export?format=xlsx")
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.content), read_only=True, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    assert rows[0] == ("a",)
    assert rows[1] == ('{"x": 1}',)


# --- Fix 6: human pagination returns a full page even past the 60KB budget ---

async def test_rows_page_not_truncated_by_agent_budget(auth_client):
    big = "y" * 4000
    content = ("q\n" + "\n".join(big for _ in range(20)) + "\n").encode("utf-8")
    ds = (await upload_ready(auth_client, "wide.csv", content))[0]
    assert ds["row_count"] == 20
    r = await auth_client.get(f"/api/datasets/{ds['id']}/rows?page=1&page_size=20")
    assert r.status_code == 200
    assert len(r.json()["rows"]) == 20


# --- Fix 6 (cont.): an oversized page_size/range is row-capped, not unbounded -

async def test_rows_page_size_is_row_capped(auth_client, monkeypatch):
    from app.routers import datasets as datasets_router

    monkeypatch.setattr(datasets_router, "MAX_ROWS_PER_REQUEST", 3)
    content = ("q\n" + "\n".join(f"v{i}" for i in range(10)) + "\n").encode("utf-8")
    ds = (await upload_ready(auth_client, "many.csv", content))[0]
    assert ds["row_count"] == 10
    r = await auth_client.get(f"/api/datasets/{ds['id']}/rows?page=1&page_size=1000000")
    assert r.status_code == 200
    body = r.json()
    assert len(body["rows"]) == 3
    assert body["truncated"] is True


async def test_rows_range_is_row_capped(auth_client, monkeypatch):
    from app.routers import datasets as datasets_router

    monkeypatch.setattr(datasets_router, "MAX_ROWS_PER_REQUEST", 3)
    content = ("q\n" + "\n".join(f"v{i}" for i in range(10)) + "\n").encode("utf-8")
    ds = (await upload_ready(auth_client, "many2.csv", content))[0]
    r = await auth_client.get(f"/api/datasets/{ds['id']}/rows?start_row=2&end_row=1000000")
    assert r.status_code == 200
    body = r.json()
    assert len(body["rows"]) == 3
    assert body["truncated"] is True


# --- Gap A: xlsx 导出含 XML 控制字符 / 超大整数 → 200(剔字符/串化)，不 500 ----

async def test_export_xlsx_strips_illegal_control_chars(auth_client):
    ds = (await upload_ready(auth_client, "ctrl.csv", b"q,a\nq1,he\x07llo\n"))[0]
    r = await auth_client.get(f"/api/datasets/{ds['id']}/export?format=xlsx")
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.content), read_only=True, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    wb.close()
    assert rows[1] == ("q1", "hello")          # 控制符 \x07 被剔除，可正常导出


async def test_export_xlsx_illegal_char_in_column_name(auth_client):
    ds = (await upload_ready(auth_client, "ctrlcol.csv", b"q\x07,a\nv,w\n"))[0]
    r = await auth_client.get(f"/api/datasets/{ds['id']}/export?format=xlsx")
    assert r.status_code == 200                 # 列名带控制符也不 500


async def test_export_xlsx_huge_int_stringified(auth_client):
    big = "9" * 400
    ds = (await upload_ready(auth_client, "big.jsonl", ('{"n": ' + big + "}\n").encode()))[0]
    r = await auth_client.get(f"/api/datasets/{ds['id']}/export?format=xlsx")
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.content), read_only=True, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    wb.close()
    assert rows[1] == (big,)                    # 超 float 范围大整数串化为文本，不抛 OverflowError


# --- Gap B: xlsx 上传含 datetime/date/time → _dumps_row 串化，不 500 ----------

def _datetime_xlsx() -> bytes:
    buf = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.append(["name", "dt", "d", "t"])
    ws.append(["report", datetime.datetime(2024, 3, 15, 9, 30),
               datetime.date(2024, 3, 16), datetime.time(8, 15)])
    wb.save(buf)
    return buf.getvalue()


async def test_upload_xlsx_with_datetime_cells(auth_client):
    ds = (await upload_ready(auth_client, "dates.xlsx", _datetime_xlsx()))[0]  # 修复前 datetime→500
    body = await auth_client.get(f"/api/datasets/{ds['id']}/rows?page=1&page_size=10")
    row = body.json()["rows"][0]
    assert isinstance(row["dt"], str) and "2024-03-15" in row["dt"]   # 串化为字面量
    assert isinstance(row["d"], str) and "2024-03-16" in row["d"]


def test_dumps_row_serializes_datetime_and_keeps_nan_null():
    from app.services.dataset_store import _dumps_row

    out = _dumps_row({"dt": datetime.datetime(2024, 1, 2, 3, 4), "x": float("inf")})
    parsed = json.loads(out)
    assert parsed["x"] is None                  # 非有限数仍中和 null
    assert "2024-01-02" in parsed["dt"]         # datetime 串化


# --- Gap C: 数据集 csv/xlsx 导出文件用后即焚（BackgroundTask 回收）----------

async def test_export_csv_file_is_reclaimed(auth_client):
    ds = (await upload_ready(auth_client, "leak.csv", b"a,b\n1,2\n"))[0]
    exports = settings.data_dir / "exports"
    before = set(exports.glob("*")) if exports.exists() else set()
    r = await auth_client.get(f"/api/datasets/{ds['id']}/export?format=csv")
    assert r.status_code == 200
    after = set(exports.glob("*")) if exports.exists() else set()
    assert after == before                      # 响应送达后导出文件被回收，零残留


async def test_export_xlsx_file_is_reclaimed(auth_client):
    ds = (await upload_ready(auth_client, "leak2.csv", b"a,b\n1,2\n"))[0]
    exports = settings.data_dir / "exports"
    before = set(exports.glob("*")) if exports.exists() else set()
    r = await auth_client.get(f"/api/datasets/{ds['id']}/export?format=xlsx")
    assert r.status_code == 200
    after = set(exports.glob("*")) if exports.exists() else set()
    assert after == before


# --- Gap D: csv 导出嵌套 dict/list → 合法 JSON（与 xlsx 一致、可往返）--------

async def test_export_csv_stringifies_nested_cells(auth_client):
    ds = (await upload_ready(auth_client, "nested.jsonl", b'{"a": {"x": 1}, "b": [1,2,3]}\n'))[0]
    r = await auth_client.get(f"/api/datasets/{ds['id']}/export?format=csv")
    assert r.status_code == 200
    rows = list(_csv.reader(StringIO(r.text)))
    assert rows[0] == ["a", "b"]
    assert json.loads(rows[1][0]) == {"x": 1}   # 合法 JSON，非 Python repr
    assert json.loads(rows[1][1]) == [1, 2, 3]


# --- Gap E: 范围读含表头时，表头伪行不占 max_rows 名额（不静默丢末行）------

async def test_range_header_does_not_consume_row_cap(tmp_path, session_factory):
    uid = await _user(session_factory, "range_header_user")
    src = tmp_path / "r.csv"
    src.write_text("q\n" + "\n".join(f"v{i}" for i in range(3)) + "\n", encoding="utf-8")
    async with session_factory() as s:
        ds = (await create_dataset_from_upload(
            s, user_id=uid, filename="r.csv", source_path=src,
            data_dir=settings.data_dir, shard_size=10))[0]
        page = await read_dataset_range(
            s, ds, data_dir=settings.data_dir, start_row=1, end_row=1000,
            max_rows=3)                          # 行数顶=3，数据正好 3 行
    data_rows = [r for r in page["rows"] if r.get("__row_type") != "header"]
    assert [r["q"] for r in data_rows] == ["v0", "v1", "v2"]   # 末行 v2 不被表头伪行挤掉
    assert page["truncated"] is False
