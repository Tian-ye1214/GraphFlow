"""大文件优化（分片存储 / 版本 CRUD / 导出 / 分页）线上活体测试。

打真实运行的 HTTP 服务（127.0.0.1:8000），全程用一次性 smoke 用户（登录即建），
覆盖 review 8 项修复 + 广覆盖回归 + 跨租户隔离，结束删净所有 smoke 数据回基线。
绝不碰真实用户数据，绝不输出任何密钥。

用法（backend 目录）：
    PYTHONIOENCODING=utf-8 PYTHONPATH=. .venv/Scripts/python.exe tools/large_dataset_live.py
"""
import datetime
import io
import sys
import zipfile

import httpx
from openpyxl import Workbook, load_workbook

from app.config import settings

BASE = "http://127.0.0.1:8000/api"
DATASETS_DIR = settings.data_dir / "datasets"

_results: list[tuple[bool, str]] = []


def check(ok: bool, label: str):
    _results.append((bool(ok), label))
    print(f"  {'PASS' if ok else 'FAIL'}  {label}")


def client(username: str) -> httpx.Client:
    c = httpx.Client(base_url=BASE, timeout=60.0, trust_env=False)
    r = c.post("/auth/login", json={"username": username})
    r.raise_for_status()
    return c


def upload(c: httpx.Client, name: str, content: bytes) -> httpx.Response:
    return c.post("/datasets/upload",
                  files=[("files", (name, content, "application/octet-stream"))])


def shard_dirs_for(ds_id: int) -> list:
    return [p for p in DATASETS_DIR.glob(f"*/{ds_id}") if p.is_dir()]


def multisheet_xlsx() -> bytes:
    buf = io.BytesIO()
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "alpha"
    ws1.append(["q", "a"])
    ws1.append(["007", "false"])           # 保真探针：不应变 7 / bool
    ws2 = wb.create_sheet("beta")
    ws2.append(["x"])
    ws2.append([42])
    wb.save(buf)
    return buf.getvalue()


def truncated_sheet_xlsx() -> bytes:
    buf = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.append(["a", "b"])
    ws.append([1, 2])
    wb.save(buf)
    src = zipfile.ZipFile(io.BytesIO(buf.getvalue()))
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as z:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename.endswith("sheet1.xml"):
                data = data[: len(data) // 2]   # 截一半 → 未闭合标签
            z.writestr(item, data)
    return out.getvalue()


def main():
    created: list[tuple[httpx.Client, int]] = []   # (owner_client, ds_id) 待清理

    def track(c, ds):
        created.append((c, ds["id"]))
        return ds

    alice = client("smoke_ld_alice")
    bob = client("smoke_ld_bob")

    # ---- Fix 1: 删数据集回收分片磁盘 -------------------------------------
    print("\n[Fix1] 删数据集回收分片磁盘")
    ds = upload(alice, "p.csv", b"q,a\nq1,a1\nq2,a2\n").json()[0]
    dirs = shard_dirs_for(ds["id"])
    check(bool(dirs) and any(d.glob("v*/part-*.jsonl") for d in dirs), "上传后分片落盘")
    r = alice.delete(f"/datasets/{ds['id']}")
    check(r.status_code == 200, f"删除返回 200 (实际 {r.status_code})")
    check(not shard_dirs_for(ds["id"]), "删除后分片目录已回收")

    # ---- Fix 2: 损坏 Excel → 422 ----------------------------------------
    print("\n[Fix2] 损坏 Excel → 422")
    r = upload(alice, "bad.xlsx", b"this is not a real workbook")
    check(r.status_code == 422, f"垃圾字节 .xlsx → 422 (实际 {r.status_code})")

    # ---- Fix 7: 合法 zip 但 sheet XML 截断 → 422（复审阻断）-------------
    print("\n[Fix7] sheet XML 截断 → 422（不 500）")
    r = upload(alice, "broken.xlsx", truncated_sheet_xlsx())
    check(r.status_code == 422, f"截断 sheet → 422 (实际 {r.status_code})")

    # ---- Fix 8: 句柄不锁，截断后正常 xlsx 仍能上传（WinError32 回归）----
    print("\n[Fix8] 截断 422 后正常 xlsx 仍可上传（句柄已释放）")
    r = upload(alice, "ok.xlsx", multisheet_xlsx())
    ok = r.status_code == 200
    check(ok, f"正常多 sheet xlsx → 200 (实际 {r.status_code})")
    if ok:
        sheets = r.json()
        check(len(sheets) == 2, f"多 sheet → 2 个数据集 (实际 {len(sheets)})")
        for d in sheets:
            track(alice, d)
        alpha = next((d for d in sheets if d["name"].endswith("alpha")), None)
        if alpha:
            rows = alice.get(f"/datasets/{alpha['id']}/rows?page=1&page_size=10").json()["rows"]
            check(rows and rows[0].get("q") == "007", f"Excel 保真 007 不变 7 (实际 {rows[0] if rows else None})")
            check(rows and rows[0].get("a") == "false", "Excel 保真 false 不变 bool")

    # ---- Fix 3: 超长 CSV 单元格不 500（>128KB 默认上限）----------------
    print("\n[Fix3] 超长 CSV 单元格被接受")
    big = "x" * 200_000
    r = upload(alice, "big.csv", f"q,a\nq1,{big}\n".encode("utf-8"))
    check(r.status_code == 200 and r.json()[0]["row_count"] == 1,
          f"200KB 单元格 → 200 & 1 行 (实际 {r.status_code})")
    if r.status_code == 200:
        track(alice, r.json()[0])

    # ---- Fix 4a: JSONL 上传含 Infinity/NaN → 解析期中和 null -----------
    print("\n[Fix4a] JSONL 非有限数上传 → null（解析期中和）")
    r = upload(alice, "nan.jsonl", b'{"a": Infinity, "b": NaN, "c": 1}\n')
    ok = r.status_code == 200
    check(ok, f"含 Infinity/NaN 的 jsonl → 200 (实际 {r.status_code})")
    if ok:
        ds = track(alice, r.json()[0])
        body = alice.get(f"/datasets/{ds['id']}/rows?page=1&page_size=10")
        check(body.status_code == 200, f"读取不 500 (实际 {body.status_code})")
        row = body.json()["rows"][0]
        check(row.get("a") is None and row.get("b") is None and row.get("c") == 1,
              f"非有限数读回为 null (实际 {row})")
        ex = alice.get(f"/datasets/{ds['id']}/export?format=jsonl")
        check(ex.status_code == 200 and b"Infinity" not in ex.content and b"NaN" not in ex.content,
              "jsonl 导出不含 Infinity/NaN")

    # ---- Fix 4b: CRUD add_constant_column 原始 Infinity → 写盘期中和 ----
    print("\n[Fix4b] CRUD 常量列 Infinity → null（写盘期 _dumps_row 中和）")
    base = upload(alice, "crud.csv", b"q\nq1\n").json()[0]
    track(alice, base)
    r = alice.post(
        f"/datasets/{base['id']}/versions",
        content='{"operations":[{"op":"add_constant_column","name":"c","value":Infinity}]}',
        headers={"content-type": "application/json"})
    ok = r.status_code == 200
    check(ok, f"常量列 Infinity 建版本 → 200 (实际 {r.status_code} {r.text[:120]})")
    if ok:
        newds = track(alice, r.json())
        body = alice.get(f"/datasets/{newds['id']}/rows?page=1&page_size=10")
        check(body.status_code == 200, f"读新版本不 500 (实际 {body.status_code})")
        row = body.json()["rows"][0]
        check(row.get("c") is None, f"常量 Infinity 列读回 null (实际 {row})")
        for p in newds and shard_dirs_for(newds["id"]):
            for f in p.glob("v*/part-*.jsonl"):
                txt = f.read_text(encoding="utf-8")
                check("Infinity" not in txt and "NaN" not in txt, f"分片 {f.name} 落盘无 Infinity/NaN")

    # ---- Fix 5: 嵌套单元格导出 xlsx 串化不 500 -------------------------
    print("\n[Fix5] 嵌套 dict/list 单元格导出 xlsx 串化")
    ds = upload(alice, "nested.jsonl", b'{"a": {"x": 1}, "b": [1,2,3]}\n').json()[0]
    track(alice, ds)
    r = alice.get(f"/datasets/{ds['id']}/export?format=xlsx")
    check(r.status_code == 200, f"嵌套导出 xlsx → 200 (实际 {r.status_code})")
    if r.status_code == 200:
        wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
        wb.close()
        check(rows[1][0] == '{"x": 1}', f"嵌套 dict 串化 (实际 {rows[1][0]!r})")

    # ---- Fix 6: 分页不被 agent 60KB 预算腰斩 + MAX_ROWS 行数顶 ----------
    print("\n[Fix6] 分页足额 + 超大 page_size 行数顶")
    wide = "y" * 4000
    content = ("q\n" + "\n".join(wide for _ in range(20)) + "\n").encode("utf-8")
    ds = upload(alice, "wide.csv", content).json()[0]
    track(alice, ds)
    check(ds["row_count"] == 20, f"20 宽行上传 (实际 {ds['row_count']})")
    body = alice.get(f"/datasets/{ds['id']}/rows?page=1&page_size=20").json()
    check(len(body["rows"]) == 20, f"宽行整页足额 20 行不被 60KB 腰斩 (实际 {len(body['rows'])})")

    rows6k = ("k\n" + "\n".join(f"v{i}" for i in range(6000)) + "\n").encode("utf-8")
    big_ds = upload(alice, "many.csv", rows6k).json()[0]
    track(alice, big_ds)
    check(big_ds["row_count"] == 6000, f"6000 行上传 (实际 {big_ds['row_count']})")
    body = alice.get(f"/datasets/{big_ds['id']}/rows?page=1&page_size=10000").json()
    check(len(body["rows"]) == 5000 and body["truncated"] is True,
          f"超大 page_size 顶在 5000 且 truncated=True (实际 {len(body['rows'])}/{body['truncated']})")
    body = alice.get(f"/datasets/{big_ds['id']}/rows?start_row=2&end_row=1000000").json()
    check(len(body["rows"]) == 5000 and body["truncated"] is True,
          f"超大范围顶在 5000 且 truncated=True (实际 {len(body['rows'])}/{body['truncated']})")

    # ==== 复审追出的后续缺口（重启后应全绿）==============================

    # Gap A: xlsx 导出含 XML 控制字符 / 超大整数 → 200(剔/串化)，不 500
    print("\n[GapA] xlsx 导出含控制字符/超大整数不 500")
    ds = upload(alice, "ctrl.csv", b"q,a\nq1,he\x07llo\n").json()[0]
    track(alice, ds)
    r = alice.get(f"/datasets/{ds['id']}/export?format=xlsx")
    check(r.status_code == 200, f"含 \\x07 控制字符导出 xlsx → 200 (实际 {r.status_code})")
    if r.status_code == 200:
        wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
        rows = list(wb.active.iter_rows(values_only=True)); wb.close()
        check(rows[1] == ("q1", "hello"), f"控制符被剔除 (实际 {rows[1] if len(rows) > 1 else None})")
    ds = upload(alice, "ctrlcol.csv", b"q\x07,a\nv,w\n").json()[0]
    track(alice, ds)
    check(alice.get(f"/datasets/{ds['id']}/export?format=xlsx").status_code == 200, "列名带控制符导出 xlsx → 200")
    big = "9" * 400
    ds = upload(alice, "big.jsonl", ('{"n": ' + big + "}\n").encode()).json()[0]
    track(alice, ds)
    r = alice.get(f"/datasets/{ds['id']}/export?format=xlsx")
    check(r.status_code == 200, f"400 位大整数导出 xlsx → 200 (实际 {r.status_code})")

    # Gap B: xlsx 上传含 datetime/date/time → 200，串化为字面量
    print("\n[GapB] xlsx 上传含 datetime 不 500")
    buf = io.BytesIO()
    wb = Workbook(); ws = wb.active
    ws.append(["name", "dt", "d", "t"])
    ws.append(["report", datetime.datetime(2024, 3, 15, 9, 30),
               datetime.date(2024, 3, 16), datetime.time(8, 15)])
    wb.save(buf)
    r = upload(alice, "dates.xlsx", buf.getvalue())
    check(r.status_code == 200, f"datetime xlsx 上传 → 200 (实际 {r.status_code})")
    if r.status_code == 200:
        ds = track(alice, r.json()[0])
        row = alice.get(f"/datasets/{ds['id']}/rows?page=1&page_size=10").json()["rows"][0]
        check(isinstance(row["dt"], str) and "2024-03-15" in row["dt"], f"datetime 串化 (实际 {row.get('dt')!r})")
        check(isinstance(row["d"], str) and "2024-03-16" in row["d"], f"date 串化 (实际 {row.get('d')!r})")

    # Gap C: 数据集导出文件用后即焚（exports/ 零残留）
    print("\n[GapC] 导出文件用后即焚")
    exports = settings.data_dir / "exports"
    ds = upload(alice, "leak.csv", b"a,b\n1,2\n").json()[0]
    track(alice, ds)
    before = set(exports.glob("*")) if exports.exists() else set()
    alice.get(f"/datasets/{ds['id']}/export?format=csv")
    alice.get(f"/datasets/{ds['id']}/export?format=xlsx")
    after = set(exports.glob("*")) if exports.exists() else set()
    check(after == before, f"csv+xlsx 导出后 exports/ 零残留 (新增 {len(after - before)})")

    # Gap D: csv 导出嵌套 dict/list → 合法 JSON（与 xlsx 一致）
    print("\n[GapD] csv 嵌套单元格串成合法 JSON")
    ds = upload(alice, "nested2.jsonl", b'{"a": {"x": 1}, "b": [1,2,3]}\n').json()[0]
    track(alice, ds)
    r = alice.get(f"/datasets/{ds['id']}/export?format=csv")
    check(r.status_code == 200, f"嵌套导出 csv → 200 (实际 {r.status_code})")
    if r.status_code == 200:
        import csv as _csv
        import json as _json
        rows = list(_csv.reader(io.StringIO(r.text)))
        ok = rows[0] == ["a", "b"] and _json.loads(rows[1][0]) == {"x": 1} and _json.loads(rows[1][1]) == [1, 2, 3]
        check(ok, f"csv 嵌套单元格可往返解析 (实际 {rows[1] if len(rows) > 1 else None})")

    # Gap E: 范围读 start_row=1 含表头，命中 5000 顶不丢末行
    print("\n[GapE] 范围读表头伪行不占行数顶")
    rows5k = ("q\n" + "\n".join(f"v{i}" for i in range(5000)) + "\n").encode("utf-8")
    ds = upload(alice, "hdr5k.csv", rows5k).json()[0]
    track(alice, ds)
    body = alice.get(f"/datasets/{ds['id']}/rows?start_row=1&end_row=5001").json()
    data = [r for r in body["rows"] if r.get("__row_type") != "header"]
    check(len(data) == 5000 and data[-1]["q"] == "v4999",
          f"含表头范围读满 5000 数据行不丢末行 (实际 {len(data)}, 末行 {data[-1].get('q') if data else None})")

    # ---- 广覆盖回归：多格式上传 + 导出往返 + 版本 CRUD 链 ---------------
    print("\n[回归] 多格式上传 + 导出往返")
    fmts = {
        "r.csv": b"a,b\n1,2\n3,4\n",
        "r.jsonl": b'{"a":1,"b":2}\n{"a":3,"b":4}\n',
        "r.json": b'[{"a":1,"b":2},{"a":3,"b":4}]',
    }
    for name, content in fmts.items():
        d = upload(alice, name, content).json()[0]
        track(alice, d)
        check(d["row_count"] == 2, f"{name} → 2 行 (实际 {d['row_count']})")
        for fmt in ("jsonl", "csv", "xlsx", "original"):
            ex = alice.get(f"/datasets/{d['id']}/export?format={fmt}")
            check(ex.status_code == 200, f"{name} 导出 {fmt} → 200 (实际 {ex.status_code})")

    print("\n[回归] 版本化 CRUD 链")
    d = upload(alice, "crudchain.csv", b"q,a\nq1,a1\nq2,a2\nq3,a3\n").json()[0]
    track(alice, d)
    # 行号均为文件可见 1-based（CSV 第 1 行=表头，q1/q2/q3 在文件行 2/3/4）
    steps = [
        ({"op": "rename_column", "from": "a", "to": "ans"}, 3, ["q", "ans"]),
        ({"op": "add_constant_column", "name": "tag", "value": "T"}, 3, ["q", "ans", "tag"]),
        ({"op": "drop_column", "name": "tag"}, 3, ["q", "ans"]),
        ({"op": "delete_rows", "start_row": 3, "end_row": 3}, 2, ["q", "ans"]),         # 删 q2(文件行3)
        ({"op": "insert_rows", "before_row": 4, "rows": [{"q": "q9", "ans": "a9"}]}, 3, ["q", "ans"]),  # 末尾追加
    ]
    cur = d
    for op, exp_rows, exp_cols in steps:
        r = alice.post(f"/datasets/{cur['id']}/versions", json={"operations": [op]})
        ok = r.status_code == 200
        check(ok, f"CRUD {op['op']} → 200 (实际 {r.status_code} {r.text[:100] if not ok else ''})")
        if ok:
            cur = track(alice, r.json())
            check(cur["row_count"] == exp_rows, f"  {op['op']} 后 {exp_rows} 行 (实际 {cur['row_count']})")
            check(cur["columns"] == exp_cols, f"  {op['op']} 后列 {exp_cols} (实际 {cur['columns']})")

    # ---- 跨租户隔离：bob 不能看/删 alice 的数据集 ---------------------
    print("\n[隔离] bob 不能访问 alice 的数据集")
    victim = upload(alice, "secret.csv", b"s\n1\n").json()[0]
    track(alice, victim)
    check(bob.get(f"/datasets/{victim['id']}/rows").status_code == 404, "bob GET rows → 404")
    check(bob.get(f"/datasets/{victim['id']}/export").status_code == 404, "bob 导出 → 404")
    check(bob.post(f"/datasets/{victim['id']}/versions", json={"operations": []}).status_code == 404,
          "bob 建版本 → 404")
    check(bob.delete(f"/datasets/{victim['id']}").status_code == 404, "bob 删除 → 404")
    check(alice.get(f"/datasets/{victim['id']}/rows").status_code == 200, "alice 自己仍可读（受害完好）")

    # ---- 参数校验：非法 page / range → 422 -----------------------------
    print("\n[校验] 非法分页参数 → 422")
    check(alice.get(f"/datasets/{victim['id']}/rows?page=0").status_code == 422, "page=0 → 422")
    check(alice.get(f"/datasets/{victim['id']}/rows?start_row=5&end_row=2").status_code == 422,
          "end<start → 422")

    # ---- 清理：删净所有 smoke 数据集（顺带复证 fix1 回收）-------------
    print("\n[清理] 删净所有 smoke 数据集回基线")
    seen = set()
    leftover_dirs = []
    for c, ds_id in created:
        if ds_id in seen:
            continue
        seen.add(ds_id)
        c.delete(f"/datasets/{ds_id}")
    for c, ds_id in created:
        if shard_dirs_for(ds_id):
            leftover_dirs.append(ds_id)
    check(not leftover_dirs, f"所有 smoke 分片目录已回收 (残留 {leftover_dirs})")
    a_left = len(alice.get("/datasets").json())
    b_left = len(bob.get("/datasets").json())
    check(a_left == 0 and b_left == 0, f"两 smoke 用户数据集清零 (alice={a_left} bob={b_left})")

    # ---- 汇总 ----------------------------------------------------------
    passed = sum(1 for ok, _ in _results if ok)
    failed = [label for ok, label in _results if not ok]
    print(f"\n{'='*60}\n活体结果：{passed}/{len(_results)} PASS")
    if failed:
        print("失败项：")
        for label in failed:
            print(f"  - {label}")
        sys.exit(1)
    print("全部通过。")


if __name__ == "__main__":
    main()
