import csv
import json
import math
import zlib
from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.engine.columns import ordered_union
from app.models import Dataset, DatasetRow

SHARD_SIZE = 100_000
MAX_AGENT_ROWS = 500
MAX_AGENT_CHARS = 60_000
MAX_JSON_NESTING = 1000
EXCEL_MAX_DATA_ROWS_PER_SHEET = 1_048_575
CSV_FIELD_LIMIT = 16 * 1024 * 1024     # 单元格上限放宽到 16MB(默认 128KB 会让长文档行误抛 csv.Error→500)
# 解析不可信 Excel 二进制可能抛的异常族(损坏/伪造/老 .xls)，统一归一为 ValueError → 上传边界转 422 而非 500。
# SyntaxError 覆盖「合法 zip 但 sheet XML 截断/畸形」在惰性行迭代时抛的 xml.etree ParseError 与
# lxml XMLSyntaxError(二者都是 SyntaxError 子类)；此路径无 eval/compile，不会误吞真实语法错误。
_EXCEL_PARSE_ERRORS = (BadZipFile, InvalidFileException, KeyError, OSError, zlib.error, SyntaxError)


def dataset_root(data_dir: Path, user_id: int, dataset_id: int, version: int) -> Path:
    return data_dir / "datasets" / str(user_id) / str(dataset_id) / f"v{version}"


def load_manifest(ds: Dataset) -> dict:
    return json.loads(ds.manifest_json or "{}")


def dump_manifest(manifest: dict) -> str:
    return json.dumps(manifest, ensure_ascii=False, separators=(",", ":"))


def _dumps_row(row: dict) -> str:
    """canonical 分片行序列化单切口：allow_nan=False 禁止非标准 JSON token；
    回退分支同时处理两类非 JSON 原生值：
      - NaN/Infinity(CRUD 经 Starlette json.loads 放进来) → _finite_only 中和 null(与读侧 parse_constant 一致)；
      - datetime/date/time/timedelta(openpyxl read_only 读 Excel 日期/时间格返回原生对象) → default=str 串化 ISO 字面量。
    json.dumps 对前者抛 ValueError、后者抛 TypeError，故两类都接；保证落盘永远是合法可读回的 JSON。"""
    try:
        return json.dumps(row, ensure_ascii=False, allow_nan=False, separators=(",", ":"))
    except (ValueError, TypeError):
        return json.dumps(_finite_only(row), ensure_ascii=False, allow_nan=False,
                          separators=(",", ":"), default=str)


def _finite_only(obj):
    if isinstance(obj, float) and not math.isfinite(obj):
        return None
    if isinstance(obj, dict):
        return {k: _finite_only(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_finite_only(v) for v in obj]
    return obj


def visible_total(row_count: int, header_row: int | None) -> int:
    return row_count + (1 if header_row is not None else 0)


@dataclass
class ParseUnit:
    """一个待建数据集的解析规格：结构探测阶段产出，后台批量写分片阶段消费。"""
    name: str
    columns: list[str]               # 已知表头(jsonl/json 无表头→[]，写分片时由 seen_columns 补全)
    header_row: int | None
    data_start_row: int
    original_format: str             # 存储/导出用的格式标识：csv / jsonl / xlsx
    reader: str                      # 源读取方式：csv / jsonl / json / xlsx
    sheet_index: int | None = None   # xlsx 专用：源工作簿里第几个 sheet


def detect_upload_structure(filename: str, source_path: Path) -> list[ParseUnit]:
    """廉价结构探测：只读表头/枚举 sheet，不读全量数据。供上传端点同步建占位数据集。"""
    suffix = Path(filename).suffix.lower()
    stem = Path(filename).stem
    if suffix == ".csv":
        return [ParseUnit(stem, _csv_header(source_path), 1, 2, "csv", "csv")]
    if suffix == ".jsonl":
        return [ParseUnit(stem, [], None, 1, "jsonl", "jsonl")]
    if suffix == ".json":
        return [ParseUnit(stem, [], None, 1, "jsonl", "json")]
    if suffix in (".xlsx", ".xls"):
        return _detect_excel_units(stem, source_path)
    raise ValueError(f"不支持的文件格式: {suffix}")


def _csv_header(path: Path) -> list[str]:
    csv.field_size_limit(CSV_FIELD_LIMIT)
    encoding = _detect_text_encoding(path)
    with path.open(encoding=encoding, newline="") as fh:
        reader = csv.reader((line.replace("\x00", "") for line in fh))
        try:
            return _disambiguate_columns(next(reader))
        except StopIteration:
            return []
        except csv.Error as exc:
            raise ValueError(f"CSV 解析失败: {exc}") from exc


def _detect_excel_units(stem: str, source_path: Path) -> list[ParseUnit]:
    """枚举 sheet + 读每 sheet 表头与首数据行；无表头或无数据行的 sheet 不建数据集(与原行为一致)。"""
    wb = None
    try:
        wb = load_workbook(source_path, read_only=True, data_only=True)
        found: list[tuple[int, str, list[str]]] = []
        for idx, ws in enumerate(wb.worksheets):
            rows = ws.iter_rows(values_only=True)
            try:
                raw_header = next(rows)
            except StopIteration:
                continue
            headers = _disambiguate_columns([_cell_to_str(v) for v in raw_header])
            if next(_rows_from_values(headers, rows), None) is None:
                continue                       # 仅表头无数据行 → 跳过
            found.append((idx, ws.title, headers))
        single = len(found) == 1
        return [ParseUnit(
            name=stem if single else f"{stem}-{title}",
            columns=headers, header_row=1, data_start_row=2,
            original_format="xlsx", reader="xlsx", sheet_index=idx,
        ) for idx, title, headers in found]
    except _EXCEL_PARSE_ERRORS as exc:
        raise ValueError(f"无法解析 Excel: {exc}") from exc
    finally:
        if wb is not None:
            wb.close()


def rows_for_unit(unit: ParseUnit, source_path: Path) -> Iterable[dict]:
    """按 ParseUnit 从源文件取数据行迭代器。纯文件 IO，宜在 to_thread 内消费。"""
    if unit.reader == "csv":
        return _iter_csv_rows(source_path)
    if unit.reader == "jsonl":
        return _iter_jsonl_rows(source_path)
    if unit.reader == "json":
        return _iter_json_rows(source_path)
    if unit.reader == "xlsx":
        return _iter_excel_sheet_rows(source_path, unit.sheet_index, unit.columns)
    raise ValueError(f"未知 reader: {unit.reader}")


def _iter_excel_sheet_rows(source_path: Path, sheet_index: int, headers: list[str]):
    wb = None
    try:
        wb = load_workbook(source_path, read_only=True, data_only=True)
        rows = wb.worksheets[sheet_index].iter_rows(values_only=True)
        next(rows, None)                       # 跳过表头行
        yield from _rows_from_values(headers, rows)
    except _EXCEL_PARSE_ERRORS as exc:
        raise ValueError(f"无法解析 Excel: {exc}") from exc
    finally:
        if wb is not None:
            wb.close()


def parse_and_write_shards(
    *,
    source_path: Path,
    unit: ParseUnit,
    data_dir: Path,
    user_id: int,
    dataset_id: int,
    version: int,
    shard_size: int = SHARD_SIZE,
    progress_cb=None,
) -> tuple[dict, list[str], int]:
    """批量：按 unit 读源文件行并写分片。纯同步、纯文件 IO，供后台 to_thread 调用。"""
    manifest, columns, row_count = _write_shards(
        rows_for_unit(unit, source_path),
        data_dir=data_dir, user_id=user_id, dataset_id=dataset_id, version=version,
        shard_size=shard_size, columns=(unit.columns or None),
        header_row=unit.header_row, data_start_row=unit.data_start_row,
        progress_cb=progress_cb)
    manifest["original_format"] = unit.original_format
    return manifest, columns, row_count


async def create_dataset_from_upload(
    session: AsyncSession,
    *,
    user_id: int,
    filename: str,
    source_path: Path,
    data_dir: Path,
    shard_size: int = SHARD_SIZE,
) -> list[Dataset]:
    """同步路径(旧调用方/兼容)：探测结构→建库→批量写分片→ready。后台异步摄入见 ingest_manager。"""
    units = detect_upload_structure(filename, source_path)
    datasets: list[Dataset] = []
    for unit in units:
        ds = Dataset(
            user_id=user_id, name=unit.name, source="upload", original_filename=filename,
            original_format=unit.original_format, file_path=str(source_path), row_count=0,
            columns_json=json.dumps(unit.columns, ensure_ascii=False), status="importing",
            header_row=unit.header_row, data_start_row=unit.data_start_row,
            total_rows_including_header=0,
        )
        session.add(ds)
        await session.flush()
        manifest, final_columns, row_count = parse_and_write_shards(
            source_path=source_path, unit=unit, data_dir=data_dir, user_id=user_id,
            dataset_id=ds.id, version=ds.version, shard_size=shard_size)
        ds.manifest_json = dump_manifest(manifest)
        ds.columns_json = json.dumps(final_columns, ensure_ascii=False)
        ds.row_count = row_count
        ds.imported_rows = row_count
        ds.total_rows_including_header = visible_total(row_count, unit.header_row)
        ds.status = "ready"
        await session.commit()
        datasets.append(ds)
    return datasets


def _write_shards(
    rows: Iterable[dict],
    *,
    data_dir: Path,
    user_id: int,
    dataset_id: int,
    version: int,
    shard_size: int,
    columns: list[str] | None,
    header_row: int | None,
    data_start_row: int,
    progress_cb=None,
) -> tuple[dict, list[str], int]:
    root = dataset_root(data_dir, user_id, dataset_id, version)
    root.mkdir(parents=True, exist_ok=True)
    shards: list[dict] = []
    current = None
    row_count = 0
    shard_rows = 0
    shard_no = 0
    seen_columns = list(columns or [])

    def open_shard():
        nonlocal current, shard_no, shard_rows
        shard_no += 1
        shard_rows = 0
        path = root / f"part-{shard_no:06d}.jsonl"
        current = path.open("w", encoding="utf-8", newline="\n")
        shards.append({
            "path": path.relative_to(data_dir).as_posix(),
            "start_data_idx": row_count,
            "start_file_row": data_start_row + row_count,
            "row_count": 0,
            "columns": [],
        })

    def close_shard():
        # 在分片关闭时一次性定稿 row_count/columns，避免逐行 list(seen_columns) 复制(宽表 O(行×列))
        nonlocal current
        if current is not None:
            current.close()
            shards[-1]["row_count"] = shard_rows
            shards[-1]["columns"] = list(seen_columns)
            current = None
            if progress_cb is not None:
                progress_cb(row_count)

    try:
        for row in rows:
            if not isinstance(row, dict):
                raise ValueError("数据行必须是 JSON object")
            if current is None or shard_rows >= shard_size:
                close_shard()
                open_shard()
            for key in row:
                if key not in seen_columns:
                    seen_columns.append(key)
            current.write(_dumps_row(row) + "\n")
            shard_rows += 1
            row_count += 1
    finally:
        close_shard()
    manifest = {
        "original_format": "",
        "canonical_format": "jsonl",
        "version": version,
        "version_of_dataset_id": None,
        "header_row": header_row,
        "data_start_row": data_start_row,
        "shard_size": shard_size,
        "shards": shards,
    }
    return manifest, seen_columns, row_count


async def ensure_dataset_materialized(session: AsyncSession, ds: Dataset, data_dir: Path) -> Dataset:
    if load_manifest(ds).get("shards"):
        return ds
    recs = (await session.execute(
        select(DatasetRow).where(DatasetRow.dataset_id == ds.id).order_by(DatasetRow.idx)
    )).scalars().all()
    if not recs:
        return ds
    suffix = Path(ds.original_filename or "").suffix.lower()
    header_row = 1 if suffix in (".csv", ".xlsx", ".xls") else None
    data_start_row = 2 if header_row is not None else 1
    original_format = suffix.lstrip(".") if suffix else "jsonl"
    rows = (json.loads(r.data_json) for r in recs)
    manifest, columns, row_count = _write_shards(
        rows, data_dir=data_dir, user_id=ds.user_id, dataset_id=ds.id, version=ds.version,
        shard_size=SHARD_SIZE, columns=json.loads(ds.columns_json or "[]"),
        header_row=header_row, data_start_row=data_start_row)
    manifest["original_format"] = original_format
    ds.manifest_json = dump_manifest(manifest)
    ds.original_format = original_format
    ds.header_row = header_row
    ds.data_start_row = data_start_row
    ds.row_count = row_count
    ds.imported_rows = row_count
    ds.total_rows_including_header = visible_total(row_count, header_row)
    ds.columns_json = json.dumps(columns, ensure_ascii=False)
    ds.status = "ready"
    await session.commit()
    return ds


async def read_dataset_range(
    session: AsyncSession,
    ds: Dataset,
    *,
    data_dir: Path,
    start_row: int,
    end_row: int,
    columns: list[str] | None = None,
    max_rows: int | None = None,
    max_chars: int | None = None,
) -> dict:
    ds = await ensure_dataset_materialized(session, ds, data_dir)
    selected = _select_columns(ds, columns)
    rows: list[dict] = []
    truncated = False
    # max_chars=None → 默认 agent 预算；max_chars=0 → 关预算(人类分页/范围读取要足额返回)；其余按值。
    char_budget = MAX_AGENT_CHARS if max_chars is None else (max_chars or None)
    used_chars = 0
    data_count = 0   # 只统计数据行：表头伪行不占 max_rows/字符预算名额

    def append_row(row: dict) -> bool:
        nonlocal used_chars, truncated, data_count
        if max_rows is not None and data_count >= max_rows:
            truncated = True
            return False
        size = len(json.dumps(row, ensure_ascii=False))
        if char_budget is not None and data_count and used_chars + size > char_budget:
            truncated = True
            return False
        rows.append(row)
        used_chars += size
        data_count += 1
        return True

    # 表头伪行直接入列、不走 append_row：它极小且非数据行，不应占用 max_rows/字符预算名额
    # (否则范围读恰从表头起、数据行数=行数顶时会静默挤掉最后一条数据行)。
    if ds.header_row is not None and start_row <= ds.header_row <= end_row:
        rows.append({"__row_type": "header", "columns": selected})

    for file_row, row in _iter_dataset_rows(ds, data_dir, start_row=start_row, end_row=end_row):
        if file_row < ds.data_start_row:
            continue
        if not append_row(_project(row, selected)):
            break
    return _range_payload(ds, start_row, end_row, selected, rows, truncated)


async def iter_jsonl_lines(session: AsyncSession, ds: Dataset, data_dir: Path):
    ds = await ensure_dataset_materialized(session, ds, data_dir)
    for _, row in _iter_dataset_rows(ds, data_dir):
        yield json.dumps(row, ensure_ascii=False) + "\n"


async def write_jsonl_export(session: AsyncSession, ds: Dataset, data_dir: Path, path: Path) -> Path:
    ds = await ensure_dataset_materialized(session, ds, data_dir)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as out:
        for _, row in _iter_dataset_rows(ds, data_dir):
            out.write(json.dumps(row, ensure_ascii=False) + "\n")
    return path


async def write_csv_export(session: AsyncSession, ds: Dataset, data_dir: Path, path: Path) -> Path:
    ds = await ensure_dataset_materialized(session, ds, data_dir)
    columns = json.loads(ds.columns_json or "[]")
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=columns, extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        for _, row in _iter_dataset_rows(ds, data_dir):
            writer.writerow({col: _jsonify_nested(row.get(col, "")) for col in columns})
    return path


async def write_xlsx_export(session: AsyncSession, ds: Dataset, data_dir: Path, path: Path) -> Path:
    ds = await ensure_dataset_materialized(session, ds, data_dir)
    columns = json.loads(ds.columns_json or "[]")
    path.parent.mkdir(parents=True, exist_ok=True)
    header = [_xlsx_cell(col) for col in columns]   # 列名也可能带控制字符，同样剔除
    wb = Workbook(write_only=True)
    ws = None
    sheet_no = 0
    sheet_rows = 0
    for _, row in _iter_dataset_rows(ds, data_dir):
        if ws is None or sheet_rows >= EXCEL_MAX_DATA_ROWS_PER_SHEET:
            sheet_no += 1
            ws = wb.create_sheet("data" if sheet_no == 1 else f"data_{sheet_no}")
            ws.append(header)
            sheet_rows = 0
        ws.append([_xlsx_cell(row.get(col, "")) for col in columns])
        sheet_rows += 1
    if ws is None:
        ws = wb.create_sheet("data")
        ws.append(header)
    wb.save(path)
    return path


def _range_payload(ds: Dataset, start_row: int, end_row: int, columns: list[str],
                   rows: list[dict], truncated: bool) -> dict:
    return {
        "dataset_id": ds.id,
        "total": ds.row_count,
        "total_rows": ds.row_count,
        "total_rows_including_header": ds.total_rows_including_header,
        "header_row": ds.header_row,
        "data_start_row": ds.data_start_row,
        "start_row": start_row,
        "end_row": end_row,
        "columns": columns,
        "rows": rows,
        "truncated": truncated,
    }


def _iter_dataset_rows(ds: Dataset, data_dir: Path, *,
                       start_row: int | None = None,
                       end_row: int | None = None):
    manifest = load_manifest(ds)
    for shard in manifest.get("shards", []):
        shard_start = int(shard["start_file_row"])
        shard_end = shard_start + int(shard["row_count"]) - 1
        if start_row is not None and shard_end < start_row:
            continue
        if end_row is not None and shard_start > end_row:
            continue
        path = Path(shard["path"])
        if not path.is_absolute():
            path = data_dir / path
        with path.open(encoding="utf-8") as fh:
            for i, line in enumerate(fh):
                file_row = shard_start + i
                if start_row is not None and file_row < start_row:
                    continue
                if end_row is not None and file_row > end_row:
                    break
                if line.strip():
                    yield file_row, json.loads(line)


def _iter_csv_rows(path: Path):
    csv.field_size_limit(CSV_FIELD_LIMIT)
    encoding = _detect_text_encoding(path)
    with path.open(encoding=encoding, newline="") as fh:
        reader = csv.reader((line.replace("\x00", "") for line in fh))
        try:
            try:
                raw_headers = next(reader)
            except StopIteration:
                return
            headers = _disambiguate_columns(raw_headers)
            for values in reader:
                yield {
                    header: values[i] if i < len(values) and values[i] is not None else ""
                    for i, header in enumerate(headers)
                }
        except csv.Error as exc:                       # 超长字段/畸形引号 → ValueError → 422 而非 500
            raise ValueError(f"CSV 解析失败: {exc}") from exc


def _detect_text_encoding(path: Path) -> str:
    with path.open("rb") as fh:
        head = fh.read(64 * 1024)
    try:
        head.decode("utf-8-sig")
        return "utf-8-sig"
    except UnicodeDecodeError:
        return "gbk"


def _disambiguate_columns(headers: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    out: list[str] = []
    for header in headers:
        base = "" if header is None else str(header)
        if base not in counts:
            counts[base] = 0
            out.append(base)
            continue
        counts[base] += 1
        candidate = f"{base}.{counts[base]}"
        while candidate in counts:
            counts[base] += 1
            candidate = f"{base}.{counts[base]}"
        counts[candidate] = 0
        out.append(candidate)
    return out


def _iter_jsonl_rows(path: Path):
    with path.open(encoding="utf-8-sig") as fh:
        for line in fh:
            if line.strip():
                obj = _loads_json(line)
                if not isinstance(obj, dict):
                    raise ValueError("JSONL 每行必须是 JSON object")
                yield obj


def _iter_json_rows(path: Path):
    data = _loads_json(path.read_text(encoding="utf-8-sig"))
    rows = data if isinstance(data, list) else [data]
    for row in rows:
        if not isinstance(row, dict):
            raise ValueError("JSON 内容必须是 object 或 object 列表")
        yield row


def _loads_json(text: str):
    try:
        obj = json.loads(text, parse_constant=lambda _v: None)
    except RecursionError as exc:
        raise ValueError(f"JSON nesting too deep: {exc}") from exc
    _reject_deep_json(obj)
    return obj


def _reject_deep_json(obj) -> None:
    stack = [(obj, 0)]
    while stack:
        current, depth = stack.pop()
        if depth > MAX_JSON_NESTING:
            raise ValueError("JSON nesting too deep")
        if isinstance(current, dict):
            stack.extend((value, depth + 1) for value in current.values())
        elif isinstance(current, list):
            stack.extend((value, depth + 1) for value in current)


def _rows_from_values(headers: list[str], rows):
    for values in rows:
        if values is None:
            continue
        row = {headers[i]: _cell_to_value(values[i]) if i < len(values) else ""
               for i in range(len(headers))}
        if any(v != "" for v in row.values()):
            yield row


def _cell_to_str(value) -> str:
    return "" if value is None else str(value)


def _cell_to_value(value):
    return "" if value is None else value


def _jsonify_nested(value):
    """dict/list 单元格串成 JSON 文本(可往返解析)；其余原样。csv/xlsx 导出共用，保证嵌套单元格两格式一致。"""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _xlsx_cell(value):
    """xlsx(write_only)单元格容错，避免 openpyxl 在 ws.append 抛错逃逸成 500：
      - 嵌套 dict/list → JSON 文本；
      - 字符串里的 XML 非法控制字符(\\x00-\\x08 等) → 剔除(否则抛 IllegalCharacterError)；
      - 超出 float 表示范围的大整数(>2**1023) → 串化(否则 openpyxl 的 isnan 检查对其 float() 溢出抛 OverflowError)。"""
    value = _jsonify_nested(value)
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    if isinstance(value, int) and not isinstance(value, bool) and value.bit_length() > 1023:
        return str(value)
    return value


def _select_columns(ds: Dataset, columns: list[str] | None) -> list[str]:
    available = json.loads(ds.columns_json or "[]")
    if columns is None:
        return available
    return [c for c in columns if c in available]


def _project(row: dict, columns: list[str]) -> dict:
    return {col: row.get(col, "") for col in columns}
